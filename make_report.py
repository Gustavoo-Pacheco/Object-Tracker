"""
Generate an Excel report from a tracking CSV.

CSV format (no header): time, x, y, vx, vy

Usage (standalone):
    python make_report.py videos/drop_track.csv

Also called automatically by track.py after every run.
"""

import argparse
import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Column positions (1-based): time, x, y, vx, vy
COL_T  = 1
COL_X  = 2
COL_Y  = 3
COL_VX = 4
COL_VY = 5


def write_report(csv_path: str, excel_path: str | None = None) -> str:
    if not os.path.exists(csv_path):
        print(f"Error: CSV not found: {csv_path}")
        sys.exit(1)

    if excel_path is None:
        stem       = os.path.splitext(csv_path)[0]
        excel_path = stem + "_report.xlsx"

    # ── Read CSV ──────────────────────────────────────────────────────────
    with open(csv_path, newline="") as f:
        rows = list(csv.reader(f))

    if not rows:
        print("Error: CSV is empty.")
        sys.exit(1)

    # ── Workbook + Data sheet ─────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    EVEN_FILL = PatternFill("solid", fgColor="F0F4F8")
    CENTER    = Alignment(horizontal="center")

    for ci in range(1, 6):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    for ri, row in enumerate(rows, start=1):
        fill = EVEN_FILL if ri % 2 == 0 else None
        for ci, val in enumerate(row, start=1):
            try:
                val = float(val) if val != "" else None
            except ValueError:
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = CENTER
            if fill:
                cell.fill = fill

    max_row = len(rows)

    # ── Helper ────────────────────────────────────────────────────────────
    def make_scatter(x_col, y_col, x_label, y_label,
                     scatter_style="smoothMarker", w=20, h=14):
        chart = ScatterChart()
        chart.scatterStyle = scatter_style
        chart.style        = 10
        chart.title        = None
        chart.x_axis.title = x_label
        chart.y_axis.title = y_label
        chart.width        = w
        chart.height       = h
        chart.legend       = None

        xvals = Reference(ws, min_col=x_col, min_row=1, max_row=max_row)
        yvals = Reference(ws, min_col=y_col, min_row=1, max_row=max_row)
        ser   = Series(yvals, xvals)
        chart.series.append(ser)
        return chart

    # ── Trajectory sheet ──────────────────────────────────────────────────
    ws_traj = wb.create_sheet("Trajectory")
    c_traj  = make_scatter(COL_X, COL_Y, "x (m)", "y (m)",
                           scatter_style="marker", w=22, h=16)
    c_traj.series[0].marker.symbol = "circle"
    c_traj.series[0].marker.size   = 3
    c_traj.series[0].graphicalProperties.line.noFill = True
    ws_traj.add_chart(c_traj, "A1")

    # ── vx(t) sheet ───────────────────────────────────────────────────────
    ws_vx  = wb.create_sheet("vx(t)")
    c_vx   = make_scatter(COL_T, COL_VX, "time (s)", "vx (m/s)")
    c_vx.series[0].graphicalProperties.line.solidFill = "1A85FF"
    c_vx.series[0].graphicalProperties.line.width     = 18000
    c_vx.series[0].marker.symbol = "none"
    ws_vx.add_chart(c_vx, "A1")

    # ── vy(t) sheet ───────────────────────────────────────────────────────
    ws_vy  = wb.create_sheet("vy(t)")
    c_vy   = make_scatter(COL_T, COL_VY, "time (s)", "vy (m/s)")
    c_vy.series[0].graphicalProperties.line.solidFill = "D41159"
    c_vy.series[0].graphicalProperties.line.width     = 18000
    c_vy.series[0].marker.symbol = "none"
    ws_vy.add_chart(c_vy, "A1")

    # ── Save ──────────────────────────────────────────────────────────────
    wb.save(excel_path)
    print(f"  Excel report     : {excel_path}")
    return excel_path


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Build Excel report from tracking CSV")
    ap.add_argument("csv",   help="Path to _track.csv produced by track.py")
    ap.add_argument("--out", default=None, help="Output .xlsx path (default: alongside CSV)")
    args = ap.parse_args()
    write_report(args.csv, args.out)
